# pylint: disable=C0114

import os
import re
from natsort import natsorted
from openpyxl import Workbook
from openpyxl.styles import PatternFill

committee_dict = {
    "과학기술정보방송통신위원회": "과학",
    "교육위원회": "교육",
    "국방위원회": "국방",
    "국토교통위원회": "국토",
    "국회운영위원회": "국회",
    "기획재정위원회": "기획",
    "농림축산식품해양수산위원회": "농림",
    "문화체육관광위원회": "문화",
    "법제사법위원회": "법제",
    "보건복지위원회": "보건",
    "산업통상자원중소벤처기업위원회": "산업",
    "여성가족위원회": "여성",
    "외교통일위원회": "외교",
    "정무위원회": "정무",
    "행정안전위원회": "행정",
    "환경노동위원회": "환경"
}

directory = input("경로명을 입력하세요 : ")
excel = input("엑셀 경로명을 확장자와 함께 입력하세요 : ")
wb = Workbook()
ws = wb.active

headers = ["YEAR", "AUDITTYPE_CDB", "COMMITTEE_ID", "COMMITTEE_NAME",
           "ORG_ID", "ORG_NAME", "PDF_NAME", "HWP_NAME", "PBM_NAME",
           "BOOK_NAME", "DIRECTORY_CDB", "FileName", "Path"]
header_color = PatternFill(start_color='4f81bd',
                           end_color='4f81bd', fill_type='solid')
for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=header)
    ws.cell(row=1, column=col_idx).fill = header_color


def generate_unique_book_name(existing_name, base_name):  # pylint: disable=C0116
    index = 1
    new_name = f"{base_name}_{index}"
    while new_name in existing_name:
        index += 1
        new_name = f"{base_name}_{index}"
    return new_name


existing_names = set()

for root, _, files in os.walk(directory):
    for file in natsorted(files):
        first_underscore_index = file.find('_')
        second_underscore_index = file.find('_', first_underscore_index + 1)
        matches = re.findall(r'\(([^)]+)\)', file)
        max_row = ws.max_row + 1
        if matches:
            match = matches[-1]
            if match == '2':
                match = matches[-2]
            if str(match).endswith('(주'):
                match = str(match).replace(  # pylint: disable=C0103
                    '(주', '(주)')
        else:
            match = ""  # pylint: disable=C0103

        if first_underscore_index != -1 and second_underscore_index != -1:
            cmt = file[first_underscore_index + 1:second_underscore_index]
        else:
            cmt = ""  # pylint: disable=C0103

        SHORT_CMT = ""
        for key, value in committee_dict.items():
            if cmt in key:
                SHORT_CMT = value
                break

        base_book_name = SHORT_CMT + '_' + match
        unique_book_name = generate_unique_book_name(
            existing_names, base_book_name)
        existing_names.add(unique_book_name)

        ws.cell(row=max_row, column=1, value=file[:4])
        ws.cell(row=max_row, column=4, value=cmt)
        ws.cell(row=max_row, column=6, value=match)
        ws.cell(row=max_row, column=10, value=unique_book_name)
        ws.cell(row=max_row, column=12, value=file)
        ws.cell(row=max_row, column=13, value=root)

wb.save(excel)
