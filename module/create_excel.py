"""폴더 경로를 입력받아 업로드 엑셀 파일 초기 버전을 생성합니다"""


import os
import re

from natsort import natsorted
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from module.data import short_cmt_dict, committee_dict, organization_dict


def extract_number_from_filename(filename):
    """괄호 안의 숫자를 추출하는 정규 표현식"""
    match = re.search(r'\((\d+)\)', filename)
    if match and match.group(1).isdigit():
        return int(match.group(1))
    return 1


def extract_cmt(filename) -> str:
    """파일명에서 위원회 이름 추출"""
    first_underscore_index = filename.find('_')
    second_underscore_index = filename.find(
        '_', first_underscore_index + 1)
    if first_underscore_index != -1 and second_underscore_index != -1:
        cmt = filename[first_underscore_index +
                       1:second_underscore_index]
    else:
        cmt = ""

    return cmt


def create_excel():  # pylint: disable=R0914
    """폴더 경로를 입력받아 업로드 엑셀 파일 초기 버전을 생성합니다"""
    print("\n>>>>>>엑셀 생성<<<<<<\n")
    print("-"*24)
    directory = input("경로명을 입력하세요 : ")
    excel = input("엑셀 경로명을 확장자와 함께 입력하세요 : ")
    wb = Workbook()
    ws = wb.active

    headers = ["YEAR", "AUDITTYPE_CDB", "COMMITTEE_ID", "COMMITTEE_NAME",
               "ORG_ID", "ORG_NAME", "PDF_NAME", "HWP_NAME", "PBM_NAME",
               "BOOK_NAME", "DIRECTORY_CDB", "FileName", "Path"]
    header_color = PatternFill(start_color='4f81bd',
                               end_color='4f81bd', fill_type='solid')
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=1, column=col_idx).fill = header_color

    before_filename = ''
    unique_number = 1

    for root, _, files in os.walk(directory):
        for file in natsorted([os.path.splitext(f)[0] for f in files]):
            tmp1, _ = os.path.splitext(file)
            if tmp1 == before_filename:
                before_filename = tmp1
                continue
            before_filename = tmp1
            cmt = extract_cmt(file)

            matches = re.findall(r'\(([^)]+)\)', file)
            if matches:
                org = matches[-1]
                if org == '2':
                    org = matches[-2]
                if str(org).endswith('(주'):
                    org = str(org).replace(  # pylint: disable=C0103
                        '(주', '(주)')
            else:
                org = ""  # pylint: disable=C0103
            max_row = ws.max_row + 1

            short_cmt = ""
            for key, value in short_cmt_dict.items():
                if cmt in key:
                    short_cmt = value
                    break

            base_book_name = short_cmt + '_' + org
            file_without_ext, _ = os.path.splitext(file)

            ws.cell(row=max_row, column=1, value=file[:4])
            ws.cell(row=max_row, column=3,
                    value=committee_dict[cmt] if cmt in committee_dict else None)
            ws.cell(row=max_row, column=4, value=cmt)
            ws.cell(row=max_row, column=5,
                    value=organization_dict[org] if org in organization_dict else None)
            ws.cell(row=max_row, column=6, value=org)

            if ws.cell(row=max_row - 1, column=6).value == ws.cell(row=max_row, column=6).value:
                unique_number += 1
            else:
                unique_number = 1
            unique_number_str = f"{unique_number:02}"
            unique_book_name = base_book_name + '_' + unique_number_str
            ws.cell(row=max_row, column=10, value=unique_book_name)
            ws.cell(row=max_row, column=12, value=file_without_ext)
            ws.cell(row=max_row, column=13, value=root)

    wb.save(excel)
